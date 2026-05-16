#!/usr/bin/env python3
"""
Genera 50 lead fac-simile per la demo Mik Cosentino.
Output: lead_demo_mik.csv + lead_demo_mik.xlsx

Schema ALLINEATO al form REALE di infobusiness.com/contattaci
(verificato live il 14/05/2026 via Chrome MCP).

8 campi del form Mik (tutti * obbligatori a parte Nome/Cognome):
1. Nome
2. Cognome
3. Telefono *
4. Email *
5. Cosa fai per vivere *
6. Qual è il tuo obiettivo nei prossimi 3-6 mesi? *
7. Perché è importante questo traguardo per te? *
8. Cosa pensi ti stia ostacolando dal raggiungere questo obiettivo da solo? *

Mix realistico del pubblico Infobusiness:
- 40% aspiranti/dipendenti (vogliono cambiare vita)
- 35% freelance/professionisti
- 20% imprenditori medi
- 5% imprenditori grandi

Telefoni FITTIZI — non chiamare batch sul file generato.
"""
import csv
import random
from datetime import datetime, timedelta

random.seed(42)

NOMI_M = ["Marco", "Luca", "Andrea", "Matteo", "Alessandro", "Davide", "Stefano",
          "Francesco", "Simone", "Giuseppe", "Antonio", "Roberto", "Federico",
          "Riccardo", "Daniele", "Lorenzo", "Gabriele", "Nicola", "Paolo",
          "Fabio", "Cristian", "Emanuele", "Vincenzo", "Salvatore", "Michele",
          "Diego", "Tommaso", "Alberto", "Massimiliano"]
NOMI_F = ["Giulia", "Sara", "Chiara", "Francesca", "Martina", "Alessia",
          "Federica", "Valentina", "Elisa", "Laura", "Silvia", "Roberta",
          "Cristina", "Erika", "Veronica", "Sofia", "Arianna", "Beatrice",
          "Camilla", "Eleonora", "Anna", "Marta", "Ilaria", "Serena", "Greta"]
COGNOMI = ["Rossi", "Bianchi", "Russo", "Ferrari", "Esposito", "Romano", "Colombo",
           "Bruno", "Ricci", "Marino", "Greco", "Conti", "De Luca", "Mancini",
           "Costa", "Giordano", "Rizzo", "Lombardi", "Moretti", "Barbieri",
           "Fontana", "Santoro", "Mariani", "Rinaldi", "Caruso", "Ferrara",
           "Galli", "Martini", "Leone", "Longo", "Gentile", "Martinelli",
           "Vitale", "Lombardo", "Serra", "Coppola", "De Santis", "D'Angelo",
           "Marchetti", "Parisi", "Villa", "Conte", "Ferraro", "Fabbri",
           "Bianco", "Marini", "Grassi", "Valentini", "Messina", "Sala",
           "Rocca", "Pucci", "Ramondetta", "Spagnulo", "Bertone"]

PREFISSI_CELL = ["320", "327", "328", "329", "331", "333", "334", "335", "338",
                 "339", "340", "342", "346", "347", "348", "349", "351", "366",
                 "370", "380", "388", "389", "391", "392", "393"]

DOMINI_EMAIL = ["gmail.com", "libero.it", "hotmail.it", "yahoo.it", "virgilio.it",
                "outlook.it", "tiscali.it", "alice.it", "icloud.com", "fastwebnet.it"]

SEGMENTI = [(40, "aspirante"), (35, "freelance"),
            (20, "imprenditore_medio"), (5, "imprenditore_grande")]

# CAMPO #5 — "Cosa fai per vivere?"
COSA_FAI = {
    "aspirante": [
        "Faccio l'impiegato amministrativo in una piccola azienda",
        "Lavoro come operaio specializzato in fabbrica",
        "Sono insegnante alle scuole medie",
        "Faccio il commesso in un negozio",
        "Lavoro come tecnico informatico dipendente",
        "Faccio il cassiere al supermercato",
        "Sono operatore call center in un'agenzia outbound",
        "Lavoro come segretaria in studio professionale",
        "Al momento sono disoccupato e cerco una strada",
        "Sono studente universitario all'ultimo anno",
        "Faccio il magazziniere in logistica",
        "Lavoro come cameriere serale",
    ],
    "freelance": [
        "Sono personal trainer con clienti privati",
        "Faccio il coach motivazionale 1-a-1",
        "Sono consulente marketing freelance per PMI",
        "Lavoro come estetista in proprio",
        "Faccio il fotografo freelance per matrimoni ed eventi",
        "Sono designer freelance, lavoro a progetto",
        "Sviluppo software come freelance per startup",
        "Sono avvocato con studio singolo",
        "Faccio il commercialista freelance",
        "Sono nutrizionista con clienti privati",
        "Lavoro come psicologo libero professionista",
        "Sono architetto con studio personale",
        "Faccio l'influencer e creator full-time",
        "Sono copywriter freelance per agenzie",
    ],
    "imprenditore_medio": [
        "Sono titolare di una piccola attività di servizi",
        "Ho una startup tecnologica con 3 dipendenti",
        "Sono founder di uno studio di consulenza",
        "Faccio il direttore commerciale di una PMI",
        "Sono AD di una piccola azienda manifatturiera",
        "Ho un'azienda di formazione online avviata",
        "Sono direttore marketing di un brand B2C",
        "Co-founder di un'agenzia digitale",
    ],
    "imprenditore_grande": [
        "Sono CEO di un'azienda con 50 dipendenti",
        "Founder di un gruppo aziendale multi-brand",
        "Sales Manager di un'azienda da 10M+",
        "Direttore generale di una società SPA",
        "AD di un'azienda quotata",
    ],
}

# CAMPO #6 — "Qual è il tuo obiettivo nei prossimi 3-6 mesi?"
OBIETTIVO_3_6_MESI = {
    "aspirante": [
        "Generare le prime entrate online stabili, 1000-3000€/mese",
        "Lasciare il lavoro dipendente entro fine anno",
        "Lanciare il mio primo infoprodotto che funzioni",
        "Acquisire le competenze base e iniziare a vendere",
        "Validare un'idea di business prima di mollare il posto fisso",
        "Capire la mia nicchia profittevole e partire",
    ],
    "freelance": [
        "Arrivare a 10-20.000€/mese stabili dal mio business",
        "Creare il mio primo corso online a profitto",
        "Costruire una community pagante di almeno 200 persone",
        "Lanciare un programma di gruppo ad alto valore",
        "Sistemare il funnel e portarlo in automazione",
        "Smettere di scambiare tempo per soldi",
    ],
    "imprenditore_medio": [
        "Raddoppiare il fatturato annuo",
        "Strutturare un team commerciale di 5 persone",
        "Lanciare una nuova linea di prodotto digitale",
        "Aprire un secondo livello del business",
        "Automatizzare il funnel commerciale",
    ],
    "imprenditore_grande": [
        "Espandere a livello europeo nei prossimi 12 mesi",
        "Raggiungere i 10M€ di fatturato annuo",
        "Acquisire un competitor più piccolo",
        "Digitalizzare completamente il go-to-market",
    ],
}

# CAMPO #7 — "Perché è importante questo traguardo per te?"
PERCHE_IMPORTANTE = [
    "Perché voglio finalmente avere tempo per la mia famiglia",
    "Perché sono stanco di vivere stipendio per stipendio",
    "Perché voglio dare ai miei figli un futuro diverso",
    "Perché ho 40 anni e non posso più rimandare",
    "Perché voglio dimostrare a me stesso che ne sono capace",
    "Perché voglio uscire da un lavoro che non mi rispecchia più",
    "Perché ho bisogno di indipendenza economica",
    "Perché voglio costruire qualcosa di mio prima dei 50",
    "Perché voglio essere libero di gestire il mio tempo",
    "Perché ho promesso a me stesso di cambiare entro quest'anno",
    "Perché credo nel mio potenziale e finora l'ho sprecato",
    "Perché voglio smettere di lavorare per arricchire altri",
    "Perché il mio attuale lavoro mi sta logorando",
    "Perché ho un'idea che mi gira in testa da anni e voglio realizzarla",
    "Perché voglio dare un senso più pieno alla mia vita lavorativa",
]

# CAMPO #8 — "Cosa pensi ti stia ostacolando dal raggiungere questo obiettivo da solo?"
COSA_OSTACOLA = [
    "Non ho un metodo chiaro, ho seguito tanti corsi senza concludere nulla",
    "Mi manca tempo, il lavoro attuale mi assorbe tutta la giornata",
    "Non ho ancora investito sul serio, ho paura di sbagliare",
    "Tecnicamente non so come fare (sito, automazioni, video)",
    "Non riesco a definire la nicchia profittevole",
    "La paura del giudizio degli altri mi blocca",
    "Faccio fatica a chiedere il giusto compenso per il mio lavoro",
    "Mi distraggo e perdo focus, salto da un'idea all'altra",
    "Da solo non riesco a vedermi dall'esterno e capire dove sbaglio",
    "Mi manca un team o persone con cui confrontarmi",
    "Non ho un piano d'azione concreto, vivo di entusiasmo iniziale",
    "Vado in burnout ogni volta che ci provo da solo",
    "Mi sento sopraffatto dalle troppe informazioni online",
    "Non so vendere e mi blocca il momento commerciale",
]


def pick_weighted(pairs):
    total = sum(w for w, _ in pairs)
    r = random.uniform(0, total)
    s = 0
    for w, val in pairs:
        s += w
        if r <= s:
            return val
    return pairs[-1][1]


def gen_tel():
    return f"+39{random.choice(PREFISSI_CELL)}{random.randint(1000000, 9999999)}"


def gen_email(nome_first, cognome):
    n = nome_first.lower().replace("'", "")
    c = cognome.lower().replace("'", "").replace(" ", "")
    patterns = [
        f"{n}.{c}@{random.choice(DOMINI_EMAIL)}",
        f"{n}{c}@{random.choice(DOMINI_EMAIL)}",
        f"{n}.{c}{random.randint(70, 99)}@{random.choice(DOMINI_EMAIL)}",
        f"{n}_{c}@{random.choice(DOMINI_EMAIL)}",
        f"{n[0]}.{c}@{random.choice(DOMINI_EMAIL)}",
        f"{n}{random.randint(80, 99)}@{random.choice(DOMINI_EMAIL)}",
    ]
    return random.choice(patterns)


def gen_data_iscrizione():
    today = datetime(2026, 5, 14)
    d = today - timedelta(days=random.randint(10, 800),
                          hours=random.randint(8, 22),
                          minutes=random.randint(0, 59))
    return d.strftime("%d/%m/%Y %H:%M")


def gen_lead(idx):
    seg = pick_weighted(SEGMENTI)
    nome = random.choice(NOMI_M if random.random() < 0.55 else NOMI_F)
    cognome = random.choice(COGNOMI)
    return {
        "id": f"MIK-DEMO-{idx:03d}",
        "data_iscrizione": gen_data_iscrizione(),
        "nome": nome,
        "cognome": cognome,
        "email": gen_email(nome, cognome),
        "telefono": gen_tel(),
        "cosa_fai_per_vivere": random.choice(COSA_FAI[seg]),
        "obiettivo_3_6_mesi": random.choice(OBIETTIVO_3_6_MESI[seg]),
        "perche_importante": random.choice(PERCHE_IMPORTANTE),
        "cosa_ostacola": random.choice(COSA_OSTACOLA),
        "consenso_marketing": "si",
        "consenso_whatsapp": random.choices(["si", "no"], weights=[88, 12])[0],
        "ultimo_contatto": "",
        "note_storiche": "",
    }


HEADERS = [
    "id", "data_iscrizione",
    "nome", "cognome", "email", "telefono",
    "cosa_fai_per_vivere", "obiettivo_3_6_mesi",
    "perche_importante", "cosa_ostacola",
    "consenso_marketing", "consenso_whatsapp",
    "ultimo_contatto", "note_storiche",
]

leads = [gen_lead(i) for i in range(1, 51)]

# CSV
with open("/Users/simocors/Desktop/telesales/demo_mik/lead_demo_mik.csv",
          "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=HEADERS)
    w.writeheader()
    for r in leads:
        w.writerow(r)

# XLSX
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lead Demo Mik"
    ws.append(HEADERS)
    bw = Font(bold=True, color="FFFFFF")
    dk = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for c in ws[1]:
        c.font = bw
        c.fill = dk
        c.alignment = Alignment(horizontal="left", vertical="center")
    for r in leads:
        ws.append([r[h] for h in HEADERS])
    widths = [14, 20, 14, 14, 32, 18, 45, 50, 50, 55, 9, 9, 14, 20]
    for i, w_ in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w_
    ws.freeze_panes = "A2"
    wb.save("/Users/simocors/Desktop/telesales/demo_mik/lead_demo_mik.xlsx")
except Exception as e:
    print("openpyxl non disponibile:", e)

print(f"Generati {len(leads)} lead allineati al form infobusiness.com/contattaci")
print("Primi 3 lead:")
for r in leads[:3]:
    print(f"  {r['id']} | {r['nome']} {r['cognome']} | {r['cosa_fai_per_vivere'][:50]}")
